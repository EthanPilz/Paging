from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl import load_workbook
import pandas as pd
from pathlib import Path
from datetime import datetime
import easygui
import os
import copy
from Configuration import writeConfigFile, readConfigFile, getBordersOption, getRequestnotesOption, getPrintOption, getInputStartOption, getOutputFolderOption

def getInputFile(config):

    if(getInputStartOption(config)):
        inputPath = easygui.fileopenbox(default = "~/Downloads/")
    else:
        inputPath = easygui.fileopenbox()
    return inputPath

def getOutputPath(config):
    
    if(getOutputFolderOption(config)):
        outputPath = str(Path.home() / "Downloads")
        #outputPath = "~/Downloads/"
    else:
        outputPath = easygui.diropenbox()
    return outputPath + '/PagingList_' + datetime.now().strftime("%Y-%m-%d %Ih%Mm %p") + '.xlsx'

def formatTable(config, wb, titleWidth, requestWidth, locationWidth):

    ws = wb.active

    ## Delete unneeded info from table
    ws.delete_cols(2, 14)
    ws.delete_cols(5, 4)

    ## Shift all cells to the right by one
    ws.move_range("A1:D100", rows=0, cols=1)

    ## Move call number column all the way to the left
    ws.move_range("E1:E100", rows=0, cols=-4)

    ## Save the output file
    wb.save(outputFile)

    ## Read by default 1st sheet of an excel file
    df1 = pd.read_excel(outputFile)

    ## Sort by Location then Call Number
    df2 = df1.sort_values(['Location', 'Call Number'])
    df2.to_excel(outputFile)

    ## Load back in the excel file from pandas output
    wb = load_workbook(filename = outputFile)
    ws = wb.active

    ws.delete_cols(1, 1)

    ## Wrap text
    for row in ws.iter_rows():
        for cell in row:      
            alignment = copy.copy(cell.alignment)
            alignment.wrapText=True
            alignment.vertical = "top"
            cell.alignment = alignment

    ## Make columns the correct width for visual clarity
    dims = {}
    for row in ws.rows:
        for cell in row:
            if(getBordersOption(config)):
                cell.border = Border(
                                left=Side(border_style=BORDER_THIN, color='00000000'),
                                right=Side(border_style=BORDER_THIN, color='00000000'),
                                top=Side(border_style=BORDER_THIN, color='00000000'),
                                bottom=Side(border_style=BORDER_THIN, color='00000000')
                                )
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

    ## Set column B (title) width, since it is stubborn
    ws.column_dimensions['B'].width = titleWidth

    ## Set column C (Request Notes) width if it's too large, since it makes the file unreadable
    if(getRequestnotesOption(config)):
        if (ws.column_dimensions['C'].width > 50):
            ws.column_dimensions['C'].width = requestWidth

    ## Set column D (Location) width, since it wraps poorly
    ws.column_dimensions['D'].width = locationWidth

    ## Set print area to only nonzero cells
    printArea = 'A1:D'
    printArea = printArea + str(df2.shape[0] + 1)
    ws.print_area = printArea

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = True

    ## Save the output file
    wb.save(outputFile)
    return ws



writeConfigFile()
config = readConfigFile()
inputFile = getInputFile(config)
outputFile = getOutputPath(config)

## Load file as xlsx
try:
    wb = load_workbook(filename = inputFile)  # this may raise an exception
except:
    easygui.exceptionbox()

## Run formatTable method. 
## Params: formatTable(workbook, titleWidth, requestWidth, locationWidth)
ws = formatTable(config, wb, 75, 20, 15)

## Print the file
if (getPrintOption(config)):
    os.startfile(outputFile, "print")